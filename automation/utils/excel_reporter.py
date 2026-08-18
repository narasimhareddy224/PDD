import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from automation.config.config import Config
from automation.utils.logger import get_logger

logger = get_logger("ExcelReporter")

class ExcelReporter:
    def __init__(self, results: List[Dict[str, Any]], metrics: Dict[str, Any]):
        self.results = results
        self.metrics = metrics
        self.output_dir = os.path.join(Config.REPORTS_DIR, "Excel")
        os.makedirs(self.output_dir, exist_ok=True)
        self.filepath = os.path.join(self.output_dir, "Automation_Test_Report.xlsx")

    def generate(self) -> str:
        wb = openpyxl.Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        pass_font = Font(name="Segoe UI", size=10, color="166534", bold=True)
        
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fail_font = Font(name="Segoe UI", size=10, color="991B1B", bold=True)
        
        skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        skip_font = Font(name="Segoe UI", size=10, color="92400E", bold=True)
        
        body_font = Font(name="Segoe UI", size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # -------------------------------------------------------------
        # Sheet 1: Executed Test Cases
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        headers1 = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Error Details"]
        ws1.append(headers1)
        
        for item in self.results:
            ws1.append([
                item.get("id", "N/A"),
                item.get("module", "General"),
                item.get("name", "Test Case"),
                item.get("status", "PASSED"),
                round(item.get("duration", 0.0), 3),
                item.get("priority", "P2"),
                item.get("error", "")
            ])

        # -------------------------------------------------------------
        # Sheet 2: Passed Tests
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.append(["Test ID", "Module", "Test Name", "Execution Time (s)", "Priority"])
        for item in self.results:
            if item.get("status") == "PASSED":
                ws2.append([
                    item.get("id"),
                    item.get("module"),
                    item.get("name"),
                    round(item.get("duration", 0.0), 3),
                    item.get("priority")
                ])

        # -------------------------------------------------------------
        # Sheet 3: Failed Tests
        # -------------------------------------------------------------
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.append(["Test ID", "Module", "Test Name", "Priority", "Failure Reason", "Screenshot Path"])
        for item in self.results:
            if item.get("status") == "FAILED":
                ws3.append([
                    item.get("id"),
                    item.get("module"),
                    item.get("name"),
                    item.get("priority"),
                    item.get("error", "Assertion Failure"),
                    item.get("screenshot", "N/A")
                ])

        # -------------------------------------------------------------
        # Sheet 4: Skipped Tests
        # -------------------------------------------------------------
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.append(["Test ID", "Module", "Test Name", "Skip Reason"])
        for item in self.results:
            if item.get("status") == "SKIPPED":
                ws4.append([
                    item.get("id"),
                    item.get("module"),
                    item.get("name"),
                    item.get("skip_reason", "Condition not met")
                ])

        # -------------------------------------------------------------
        # Sheet 5: Execution Metrics
        # -------------------------------------------------------------
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.append(["Metric", "Value"])
        metrics_data = [
            ["Total Test Cases", self.metrics.get("total", len(self.results))],
            ["Passed Test Cases", self.metrics.get("passed", 0)],
            ["Failed Test Cases", self.metrics.get("failed", 0)],
            ["Skipped Test Cases", self.metrics.get("skipped", 0)],
            ["Pass Percentage (%)", f"{self.metrics.get('pass_rate', 0.0):.2f}%"],
            ["Total Duration (seconds)", f"{self.metrics.get('total_duration', 0.0):.2f}s"],
            ["Execution Timestamp", self.metrics.get("timestamp", "")],
            ["Live Target URL", Config.BASE_URL],
            ["Browser Environment", f"{Config.BROWSER} (Headless: {Config.HEADLESS})"],
        ]
        for row in metrics_data:
            ws5.append(row)

        # -------------------------------------------------------------
        # Sheet 6: Defect Summary
        # -------------------------------------------------------------
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.append(["Defect ID", "Associated Test ID", "Module", "Severity", "Root Cause Summary"])
        defects = [item for item in self.results if item.get("status") == "FAILED"]
        for idx, item in enumerate(defects, start=1):
            ws6.append([
                f"BUG-{idx:03d}",
                item.get("id"),
                item.get("module"),
                "High" if item.get("priority") == "P1" else "Medium",
                str(item.get("error", "Element not found or timed out"))[:150]
            ])

        # Apply Styling across all sheets
        for sheet in wb.worksheets:
            # Header styling
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Row styling and auto width
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = body_font
                    cell.border = thin_border
                    # Color status
                    if str(cell.value) == "PASSED":
                        cell.fill = pass_fill
                        cell.font = pass_font
                    elif str(cell.value) == "FAILED":
                        cell.fill = fail_fill
                        cell.font = fail_font
                    elif str(cell.value) == "SKIPPED":
                        cell.fill = skip_fill
                        cell.font = skip_font

            # Adjust column dimensions
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 50)

        wb.save(self.filepath)
        
        # Also create standalone filtered workbooks for artifacts
        self._export_standalone_workbooks(wb)
        
        logger.info(f"Excel report successfully generated at: {self.filepath}")
        return self.filepath

    def _export_standalone_workbooks(self, main_wb: openpyxl.Workbook):
        # 1. Failed tests workbook
        failed_wb = openpyxl.Workbook()
        failed_ws = failed_wb.active
        failed_ws.title = "Failed Tests"
        source_ws = main_wb["Failed Tests"]
        for row in source_ws.iter_rows(values_only=True):
            failed_ws.append(list(row))
        failed_wb.save(os.path.join(self.output_dir, "Failed_Test_Cases.xlsx"))

        # 2. Passed tests workbook
        passed_wb = openpyxl.Workbook()
        passed_ws = passed_wb.active
        passed_ws.title = "Passed Tests"
        source_ws = main_wb["Passed Tests"]
        for row in source_ws.iter_rows(values_only=True):
            passed_ws.append(list(row))
        passed_wb.save(os.path.join(self.output_dir, "Passed_Test_Cases.xlsx"))

        # 3. Summary Report workbook
        summary_wb = openpyxl.Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = "Summary"
        source_ws = main_wb["Execution Metrics"]
        for row in source_ws.iter_rows(values_only=True):
            summary_ws.append(list(row))
        summary_wb.save(os.path.join(self.output_dir, "Summary_Report.xlsx"))
